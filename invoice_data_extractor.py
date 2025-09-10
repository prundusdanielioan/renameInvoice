#!/usr/bin/env python3
"""
Invoice Data Extractor - Program pentru extragerea datelor complete din facturi PDF
Extrage: numele companiei, data emitere, data scadenta, total plata, total TVA
"""

import os
import sys
import re
import argparse
import shutil
from pathlib import Path
import PyPDF2
from typing import List, Optional, Dict, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class InvoiceDataExtractor:
    def __init__(self, input_folder: str, dry_run: bool = False):
        """
        Inițializează Invoice Data Extractor
        
        Args:
            input_folder: Folderul cu PDF-urile de procesat
            dry_run: Dacă True, doar afișează ce ar face fără să redenumească
        """
        self.input_folder = Path(input_folder)
        self.dry_run = dry_run
        self.processed_files = []
        self.errors = []
        
        if not self.input_folder.exists():
            raise FileNotFoundError(f"Folderul {input_folder} nu există")
    
    def extract_text_pypdf2(self, pdf_path: Path) -> str:
        """Extrage text din PDF folosind PyPDF2"""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            print(f"Eroare la extragerea textului cu PyPDF2 din {pdf_path}: {e}")
            return ""
    
    def extract_text(self, pdf_path: Path) -> str:
        """Extrage text din PDF folosind PyPDF2"""
        return self.extract_text_pypdf2(pdf_path)
    
    def extract_company_name(self, text: str) -> Optional[str]:
        """Extrage numele companiei din text"""
        patterns = [
            # Caută "VANZATOR" urmat de numele companiei
            r'VANZATOR\s+([A-Za-z0-9\s\.\-&]+?)(?=\s+Nume|\n)',
            # Caută "Denumire" urmat de numele companiei
            r'Denumire\s+([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
            # Caută "Nume" urmat de spațiu și apoi numele companiei până la următoarea linie
            r'Nume\s+([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
            # Variantă mai permisivă
            r'Nume\s+([A-Za-z0-9\s\.\-&]+?)(?=\s+[A-Z]|\n|$)',
            # Caută pe linia următoare după "Nume"
            r'Nume\s*\n\s*([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                company_name = match.strip()
                company_name = re.sub(r'\s+', ' ', company_name)
                company_name = company_name.rstrip('.')
                
                # Filtrează numele prea scurte, care conțin doar cifre sau care încep cu "Nr."
                if (len(company_name) > 3 and 
                    not company_name.isdigit() and 
                    not company_name.startswith('Nr.')):
                    return company_name
        
        return None
    
    def extract_date(self, text: str, label: str) -> Optional[str]:
        """Extrage o dată specifică din text pe baza labelului"""
        # Pattern-uri pentru date în format YYYY-MM-DD sau DD-MM-YYYY
        date_patterns = [
            rf'{label}\s+(\d{{4}}-\d{{2}}-\d{{2}})',
            rf'{label}\s+(\d{{2}}-\d{{2}}-\d{{4}})',
            rf'{label}\s+(\d{{1,2}}/\d{{1,2}}/\d{{4}})',
            rf'{label}\s+(\d{{4}}/\d{{1,2}}/\d{{1,2}})',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return None
    
    def extract_total_payment(self, text: str) -> Optional[str]:
        """Extrage totalul plății din text"""
        patterns = [
            r'TOTAL\s+PLATA\s+(\d+\.?\d*)',
            r'TOTAL\s+PAYMENT\s+(\d+\.?\d*)',
            r'TOTAL\s+PLATA\s+(\d+,\d*)',
            r'TOTAL\s+PAYMENT\s+(\d+,\d*)',
            # Caută "TOTAL PLATA" urmat de număr pe linia următoare
            r'TOTAL\s+PLATA\s*\n\s*(\d+\.?\d*)',
            # Caută numere în contextul "TOTAL PLATA"
            r'(\d+\.?\d*)\s*TOTAL\s+PLATA',
            r'TOTAL\s+PLATA.*?(\d+\.?\d*)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).replace(',', '.')
        
        return None
    
    def extract_total_vat(self, text: str) -> Optional[str]:
        """Extrage totalul TVA din text"""
        patterns = [
            r'TOTAL\s+TVA\s+(\d+\.?\d*)\s*RON',
            r'TOTAL\s+VAT\s+(\d+\.?\d*)\s*RON',
            r'TOTAL\s+TVA\s+(\d+,\d*)\s*RON',
            r'TOTAL\s+VAT\s+(\d+,\d*)\s*RON',
            # Caută "TOTAL TVA" urmat de număr pe linia următoare
            r'TOTAL\s+TVA\s*\n\s*(\d+\.?\d*)\s*RON',
            # Caută numere în contextul "TOTAL TVA"
            r'(\d+\.?\d*)\s*RON\s*TOTAL\s+TVA',
            r'TOTAL\s+TVA.*?(\d+\.?\d*)\s*RON',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).replace(',', '.')
        
        return None
    
    def extract_cpv_code(self, text: str) -> Optional[str]:
        """Extrage codul CPV din text"""
        patterns = [
            r'Cod\s+CPV\s+articol\s+pentru\s+linia\s+\d+\s*:\s*([A-Z0-9]+)',
            r'Cod\s+CPV\s*:\s*([A-Z0-9]+)',
            r'CPV\s+Code\s*:\s*([A-Z0-9]+)',
            r'CPV\s*:\s*([A-Z0-9]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def extract_nc8_code(self, text: str) -> Optional[str]:
        """Extrage codul NC8 din text"""
        patterns = [
            r'Cod\s+NC8\s+articol\s+pentru\s+linia\s+\d+\s*:\s*([A-Z0-9]+)',
            r'Cod\s+NC8\s*:\s*([A-Z0-9]+)',
            r'NC8\s+Code\s*:\s*([A-Z0-9]+)',
            r'NC8\s*:\s*([A-Z0-9]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def extract_product_name(self, text: str) -> Optional[str]:
        """Extrage denumirea produsului din text"""
        patterns = [
            # Caută în tabelul cu "Nume articol/Descriere articol"
            r'Linia\s+1\s+([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
            r'Nume\s+articol/Descriere\s+articol\s*\n\s*([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
            # Caută după "Linia 1" urmat de denumirea produsului
            r'Linia\s+1\s*\n\s*([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
            # Caută în contextul tabelului de articole
            r'1\s+([A-Za-z0-9\s\.\-&]+?)(?=\n|$)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                product_name = match.group(1).strip()
                # Filtrează numele prea scurte sau care conțin doar cifre
                if len(product_name) > 2 and not product_name.isdigit():
                    return product_name
        
        return None
    
    def extract_invoice_data(self, text: str) -> Dict[str, Optional[str]]:
        """
        Extrage toate datele din text
        
        Returns:
            Dict cu toate datele extrase
        """
        data = {
            'company_name': self.extract_company_name(text),
            'issue_date': self.extract_date(text, 'Data emitere'),
            'due_date': self.extract_date(text, 'Data scadenta'),
            'total_payment': self.extract_total_payment(text),
            'total_vat': self.extract_total_vat(text),
            'cpv_code': self.extract_cpv_code(text),
            'nc8_code': self.extract_nc8_code(text),
            'product_name': self.extract_product_name(text),
        }
        
        return data
    
    def sanitize_filename(self, filename: str) -> str:
        """Curăță numele fișierului de caractere invalide"""
        # Înlocuiește caracterele invalide cu underscore
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # Elimină spațiile multiple și le înlocuiește cu underscore
        filename = re.sub(r'\s+', '_', filename)
        
        # Elimină underscore-urile multiple
        filename = re.sub(r'_+', '_', filename)
        
        # Elimină underscore-urile de la început și sfârșit
        filename = filename.strip('_')
        
        return filename
    
    def generate_new_filename(self, original_path: Path, data: Dict[str, Optional[str]]) -> Path:
        """
        Generează noul nume de fișier pe baza datelor extrase
        
        Args:
            original_path: Calea originală a fișierului
            data: Datele extrase din PDF
            
        Returns:
            Calea noului fișier
        """
        # Păstrează extensia
        extension = original_path.suffix
        
        # Construiește numele pe baza datelor disponibile
        parts = []
        
        # Adaugă numele companiei dacă există
        if data['company_name']:
            company_part = data['company_name'][:30]  # Limitează la 30 caractere
            parts.append(company_part)
        
        # Adaugă data emitere dacă există
        if data['issue_date']:
            # Convertește data în format YYYY-MM-DD dacă e necesar
            date_str = self.normalize_date(data['issue_date'])
            if date_str:
                parts.append(date_str)
        
        # Adaugă totalul plății dacă există
        if data['total_payment']:
            parts.append(f"TOTAL_{data['total_payment']}")
        
        # Dacă nu avem date suficiente, folosește numele original
        if not parts:
            return original_path
        
        # Creează noul nume
        new_name = "_".join(parts) + extension
        new_name = self.sanitize_filename(new_name)
        
        return original_path.parent / new_name
    
    def normalize_date(self, date_str: str) -> Optional[str]:
        """Normalizează data în format YYYY-MM-DD"""
        try:
            # Încearcă să parseze diferite formate de dată
            formats = [
                '%Y-%m-%d',
                '%d-%m-%Y',
                '%d/%m/%Y',
                '%Y/%m/%d',
            ]
            
            for fmt in formats:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            
            return None
        except:
            return None
    
    def process_pdf(self, pdf_path: Path) -> Tuple[bool, str, Dict[str, Optional[str]]]:
        """
        Procesează un singur PDF
        
        Args:
            pdf_path: Calea către PDF
            
        Returns:
            Tuple (success, message, extracted_data)
        """
        try:
            print(f"Procesez: {pdf_path.name}")
            
            # Extrage textul
            text = self.extract_text(pdf_path)
            if not text.strip():
                return False, "Nu s-a putut extrage text din PDF", {}
            
            # Extrage datele
            data = self.extract_invoice_data(text)
            
            # Afișează datele găsite
            print(f"  Date găsite:")
            for key, value in data.items():
                if value:
                    print(f"    {key}: {value}")
            
            # Verifică dacă am găsit cel puțin o dată
            if not any(data.values()):
                return False, "Nu s-au găsit date în PDF", data
            
            # Generează noul nume
            new_path = self.generate_new_filename(pdf_path, data)
            
            # Verifică dacă fișierul de destinație este același cu cel sursă
            if new_path == pdf_path:
                print(f"  ℹ️  Fișierul {pdf_path.name} are deja numele corect")
                return True, f"Fișierul are deja numele corect: {pdf_path.name}", data
            
            # Verifică dacă fișierul de destinație există deja
            if new_path.exists():
                print(f"  ⚠️  Fișierul {new_path.name} există deja - îl recreez")
            
            # Creează o copie cu noul nume (suprascrie dacă există)
            if not self.dry_run:
                shutil.copy2(pdf_path, new_path)
                if new_path.exists():
                    print(f"  ✅ Recreeat: {new_path.name}")
                else:
                    print(f"  ✅ Copiat în: {new_path.name}")
            else:
                if new_path.exists():
                    print(f"  [DRY RUN] Ar recreea: {new_path.name}")
                else:
                    print(f"  [DRY RUN] Ar copia în: {new_path.name}")
            
            return True, f"Succes: copiat în {new_path.name}", data
            
        except Exception as e:
            return False, f"Eroare: {str(e)}", {}
    
    def is_original_file(self, filename: str) -> bool:
        """
        Verifică dacă fișierul este original (cu nume numeric) sau creat de program
        
        Args:
            filename: Numele fișierului
            
        Returns:
            True dacă fișierul este original și trebuie procesat
        """
        # Elimină extensia
        name_without_ext = filename.replace('.pdf', '')
        
        # Verifică dacă numele conține doar cifre și eventual # la sfârșit
        if re.match(r'^\d+#?$', name_without_ext):
            return True
        
        # Verifică dacă numele conține caractere tipice de fișiere create de program
        # (conține underscore, date, sau cuvinte descriptive)
        if '_' in name_without_ext or re.search(r'\d{4}-\d{2}-\d{2}', name_without_ext):
            return False
        
        # Dacă nu se potrivește cu pattern-ul numeric, probabil este creat de program
        return False
    
    def process_folder(self) -> None:
        """Procesează toate PDF-urile din folder"""
        all_pdf_files = list(self.input_folder.glob("*.pdf"))
        
        if not all_pdf_files:
            print("Nu s-au găsit fișiere PDF în folder")
            return
        
        # Filtrează doar fișierele originale (cu nume numeric)
        pdf_files = [f for f in all_pdf_files if self.is_original_file(f.name)]
        ignored_files = [f for f in all_pdf_files if not self.is_original_file(f.name)]
        
        print(f"Găsite {len(all_pdf_files)} fișiere PDF în total")
        print(f"Fișiere originale de procesat: {len(pdf_files)}")
        print(f"Fișiere ignorate (create de program): {len(ignored_files)}")
        print(f"Folder: {self.input_folder}")
        
        if ignored_files:
            print("\nFișiere ignorate (create de program):")
            for i, pdf_file in enumerate(ignored_files[:3]):  # Afișează primele 3
                print(f"  - {pdf_file.name}")
            if len(ignored_files) > 3:
                print(f"  ... și încă {len(ignored_files) - 3} fișiere")
        
        if not pdf_files:
            print("\n❌ Nu s-au găsit fișiere originale de procesat!")
            print("💡 Fișierele originale trebuie să aibă nume numerice (ex: 5532528720#)")
            return
        
        print(f"\nFișiere originale de procesat:")
        for i, pdf_file in enumerate(pdf_files[:5]):  # Afișează primele 5
            print(f"  {i+1}. {pdf_file.name}")
        if len(pdf_files) > 5:
            print(f"  ... și încă {len(pdf_files) - 5} fișiere")
        
        if self.dry_run:
            print("\n*** MOD DRY RUN - Nu se vor copia fișierele ***")
        print("-" * 50)
        
        for pdf_file in pdf_files:
            success, message, data = self.process_pdf(pdf_file)
            
            if success:
                self.processed_files.append((pdf_file.name, message, data))
            else:
                self.errors.append((pdf_file.name, message))
        
        # Afișează rezumatul
        print("\n" + "=" * 50)
        print("REZUMAT:")
        print(f"Procesate cu succes: {len(self.processed_files)}")
        print(f"Erori: {len(self.errors)}")
        
        if self.errors:
            print("\nErori:")
            for filename, error in self.errors:
                print(f"  {filename}: {error}")
    
    def save_extracted_data(self, output_file: str = "extracted_data.txt") -> None:
        """Salvează datele extrase într-un fișier"""
        output_path = Path(output_file)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("Date extrase din PDF-uri:\n")
            f.write("=" * 40 + "\n\n")
            
            for filename, message, data in self.processed_files:
                f.write(f"Fișier: {filename}\n")
                f.write(f"Status: {message}\n")
                f.write("Date extrase:\n")
                for key, value in data.items():
                    f.write(f"  {key}: {value or 'N/A'}\n")
                f.write("-" * 30 + "\n\n")
        
        print(f"\nDatele au fost salvate în: {output_path}")
    
    def save_to_excel(self, output_file: str = "facturi_extrase.xlsx") -> None:
        """Salvează datele extrase într-un fișier Excel"""
        # Dacă nu este cale absolută, salvează în folderul de procesare
        if not Path(output_file).is_absolute():
            output_path = self.input_folder / output_file
        else:
            output_path = Path(output_file)
        
        # Creează workbook-ul
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturi Extrase"
        
        # Definește stilurile
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Definește header-urile
        headers = [
            "Fișier Original",
            "Nume Companie", 
            "Data Emitere",
            "Data Scadenta",
            "Total Plata",
            "Total TVA",
            "Denumire Produs",
            "Cod CPV",
            "Cod NC8",
            "Status",
            "Fișier Nou"
        ]
        
        # Adaugă header-urile
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adaugă datele
        for row, (filename, message, data) in enumerate(self.processed_files, 2):
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=data.get('company_name', 'N/A'))
            ws.cell(row=row, column=3, value=data.get('issue_date', 'N/A'))
            ws.cell(row=row, column=4, value=data.get('due_date', 'N/A'))
            ws.cell(row=row, column=5, value=data.get('total_payment', 'N/A'))
            ws.cell(row=row, column=6, value=data.get('total_vat', 'N/A'))
            ws.cell(row=row, column=7, value=data.get('product_name', 'N/A'))
            ws.cell(row=row, column=8, value=data.get('cpv_code', 'N/A'))
            ws.cell(row=row, column=9, value=data.get('nc8_code', 'N/A'))
            ws.cell(row=row, column=10, value="Succes")
            ws.cell(row=row, column=11, value=message.replace("Succes: copiat în ", ""))
        
        # Adaugă erorile
        for row, (filename, error) in enumerate(self.errors, len(self.processed_files) + 2):
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value="N/A")
            ws.cell(row=row, column=3, value="N/A")
            ws.cell(row=row, column=4, value="N/A")
            ws.cell(row=row, column=5, value="N/A")
            ws.cell(row=row, column=6, value="N/A")
            ws.cell(row=row, column=7, value="N/A")
            ws.cell(row=row, column=8, value="N/A")
            ws.cell(row=row, column=9, value="N/A")
            ws.cell(row=row, column=10, value="Eroare")
            ws.cell(row=row, column=11, value=error)
        
        # Ajustează lățimea coloanelor
        column_widths = [25, 30, 15, 15, 15, 15, 25, 12, 12, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Salvează fișierul
        wb.save(output_path)
        print(f"\nDatele au fost salvate în Excel: {output_path}")
        
        # Afișează statistici
        total_files = len(self.processed_files) + len(self.errors)
        success_rate = (len(self.processed_files) / total_files * 100) if total_files > 0 else 0
        
        print(f"Statistici:")
        print(f"  Total fișiere: {total_files}")
        print(f"  Procesate cu succes: {len(self.processed_files)}")
        print(f"  Erori: {len(self.errors)}")
        print(f"  Rata de succes: {success_rate:.1f}%")


def main():
    parser = argparse.ArgumentParser(
        description="Extrage date complete din facturi PDF și creează copii cu nume noi",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemple de utilizare:

1. Procesează toate PDF-urile din folderul curent:
   python invoice_data_extractor.py /path/to/folder

2. Testează fără să copieze fișierele (dry run):
   python invoice_data_extractor.py /path/to/folder --dry-run

3. Salvează datele extrase într-un fișier:
   python invoice_data_extractor.py /path/to/folder --save data.txt

4. Salvează datele extrase în Excel:
   python invoice_data_extractor.py /path/to/folder --excel facturi.xlsx

Date extrase:
- Numele companiei (după labelul "Nume")
- Data emitere (după labelul "Data emitere")
- Data scadenta (după labelul "Data scadenta")
- Total plata (după labelul "TOTAL PLATA")
- Total TVA (după labelul "TOTAL TVA")
- Denumirea produsului (din coloana "Nume articol/Descriere articol")
- Cod CPV (după "Cod CPV articol pentru linia X")
- Cod NC8 (după "Cod NC8 articol pentru linia X")

Format nume fișier: [Nume_Companie]_[Data]_TOTAL_[Valoare].pdf
        """
    )
    
    parser.add_argument(
        "folder",
        help="Folderul cu PDF-urile de procesat"
    )
    
    parser.add_argument(
        "--dry-run", "-d",
        action="store_true",
        help="Testează fără să redenumească fișierele"
    )
    
    parser.add_argument(
        "--save", "-s",
        help="Salvează datele extrase în fișierul specificat"
    )
    
    parser.add_argument(
        "--excel", "-e",
        help="Salvează datele extrase în fișierul Excel specificat"
    )
    
    args = parser.parse_args()
    
    try:
        extractor = InvoiceDataExtractor(args.folder, args.dry_run)
        extractor.process_folder()
        
        if args.save:
            extractor.save_extracted_data(args.save)
        
        if args.excel:
            extractor.save_to_excel(args.excel)
            
    except Exception as e:
        print(f"Eroare: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
